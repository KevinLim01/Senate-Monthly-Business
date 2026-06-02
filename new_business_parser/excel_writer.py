from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import BusinessRecord

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
REJECT_FILL = PatternFill("solid", fgColor="F4CCCC")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def write_excel(records: list[BusinessRecord], output_path: str | Path) -> None:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    accepted = [r for r in records if r.status == "accepted"]
    rejected = [r for r in records if r.status != "accepted"]

    accepted_df = pd.DataFrame(
        [
            {
                "Business Name": r.clean_business_name,
                "Street Address": r.mailing_street,
                "City": r.mailing_city,
                "State": r.mailing_state,
                "ZIP Code": r.mailing_zip,
                "District 24 Check": r.senate_district_24_status,
            }
            for r in accepted
        ]
    )

    rejected_df = pd.DataFrame(
        [
            {
                "Original Name": r.raw_business_name,
                "Cleaned Name": r.clean_business_name,
                "Street Address": r.mailing_street,
                "City": r.mailing_city,
                "State": r.mailing_state,
                "ZIP Code": r.mailing_zip,
                "Reason": r.reason,
                "District 24 Check": r.senate_district_24_status,
                "District 24 Reason": r.senate_district_24_reason,
                "Page": r.page,
            }
            for r in rejected
        ]
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        sheet_name = "Businesses"
        accepted_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)

        rejected_title_row = len(accepted_df) + 5
        rejected_start_row = rejected_title_row + 1
        rejected_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=rejected_start_row)

        ws = writer.book[sheet_name]
        ws.cell(row=rejected_title_row + 1, column=1).value = "REJECTED / NEEDS REVIEW"
        ws.cell(row=rejected_title_row + 1, column=1).font = Font(bold=True, size=13)
        ws.cell(row=rejected_title_row + 1, column=1).fill = REJECT_FILL

        _format_sheet(ws, rejected_start_row=rejected_start_row + 1)


def _format_sheet(ws, rejected_start_row: int) -> None:
    ws.freeze_panes = "A2"

    for cell in ws[1]:
        if cell.value:
            cell.font = Font(bold=True)
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

    for cell in ws[rejected_start_row + 1]:
        if cell.value:
            cell.font = Font(bold=True)
            cell.fill = REJECT_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = THIN_BORDER

    widths = {
        "A": 32,
        "B": 32,
        "C": 22,
        "D": 10,
        "E": 12,
        "F": 18,
        "G": 36,
        "H": 18,
        "I": 42,
        "J": 10,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 20
